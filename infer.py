import os
from PIL import Image
import torch
import json
from torchvision import transforms
from transformers import ViTForImageClassification, ViTImageProcessor
import cv2
from scipy.stats import pearsonr
from typing import List
import openpyxl
import csv
from collections import Counter

# Load preprocessor_config.json from the current directory
script_dir = os.path.dirname(os.path.abspath(__file__))
config_file_path = os.path.join(script_dir, "preprocessor_config.json")
with open(config_file_path, "r") as f:
    preprocessor_config = json.load(f)

# Model name
model_name = 'trpakov/vit-face-expression'

# Initialize ViTImageProcessor with the preprocessor_config
preprocessor = ViTImageProcessor.from_pretrained(model_name, do_rescale=False, config=preprocessor_config)
processor = ViTImageProcessor.from_pretrained(model_name, do_rescale=False)

# Load ViTForImageClassification model
model = ViTForImageClassification.from_pretrained(model_name)

def image_is_valid(image_path):
    image = cv2.imread(image_path)
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')
    faces = face_cascade.detectMultiScale(gray, scaleFactor=1.1, minNeighbors=5, minSize=(30, 30))
    return len(faces) > 0

# Preprocess input image with face cropping
def preprocess_image(image_path, target_size):
    face_image = cv2.imread(image_path)
    image = Image.fromarray(face_image)
    image = image.resize((target_size, target_size))
    image = transforms.ToTensor()(image)
    inputs = preprocessor(images=image, return_tensors="pt")
    return inputs.pixel_values

# Run model on input image, return analyzed image
def run_model(image_path):
    input_image = preprocess_image(image_path, target_size=224)
    with torch.no_grad():
        outputs = model(input_image, output_hidden_states=True)
    return outputs

# Get the predicted emotion by the model
def predict_emotion(analyzed_image):
    predicted_class = torch.argmax(analyzed_image.logits).item()
    emotions = ['Angry', 'Disgust', 'Fear', 'Happy', 'Sad', 'Surprise', 'Neutral']
    return emotions[predicted_class]

# Calculate the Pearson correlation between all pairs from 3 images over desired layer
def pearson_correlation(analyzed_images, layer_num=12):
    tensors = [analyzed_image.hidden_states[layer_num] for analyzed_image in analyzed_images]
    correlation_coefficients = []

    for i in range(len(tensors)):
        for j in range(i + 1, len(tensors)):
            correlation_coefficient = pearsonr(tensors[i].flatten().tolist(), tensors[j].flatten().tolist())
            correlation_coefficients.append((correlation_coefficient))

    if len(correlation_coefficients) == 3:
        return [correlation_coefficients[2], correlation_coefficients[1], correlation_coefficients[0]]
    return correlation_coefficients

# Return all the file paths from the desired folder
def get_files_in_folder(folder_path):
    file_paths = []
    for root, dirs, files in os.walk(folder_path):
        for file in files:
            file_paths.append(os.path.join(root, file))
    return file_paths

def get_folders_in_folder(folder_path: str) -> list:
    folder_paths = []
    for item in os.listdir(folder_path):
        item_path = os.path.join(folder_path, item)
        if os.path.isdir(item_path):
            folder_paths.append((int(item.split('_')[1]), item_path))
    sorted_folder_paths = [path for _, path in sorted(folder_paths)]
    return sorted_folder_paths

def write_to_excel_triplets_result(image_folder_path: str, data: list[List[List[float]]], output_folder: str, csv_file_path: str):
    excel_file_path = os.path.join(output_folder, "triplets_result.xlsx")
    workbook = openpyxl.load_workbook(excel_file_path) if os.path.exists(excel_file_path) else openpyxl.Workbook()
    worksheet = workbook["Triplets Result"] if "Triplets Result" in workbook.sheetnames else workbook.active
    worksheet.title = "Triplets Result"

    # Add header row
    if worksheet.max_row == 1:
        header_row = ["Triplet ID", "Image1 ID", "Image2 ID", "Image3 ID", "Human 1", "Human 2", "Human 3", "Layer Number", "Model 1", "Model 2", "Model 3", "Model Score", "dominance", "human agreement"]
        for col_idx, header in enumerate(header_row, start=1):
            worksheet.cell(row=1, column=col_idx, value=header)

    with open(csv_file_path, 'r') as file:
        reader = csv.reader(file)
        next(reader)
        triplet_index = worksheet.max_row
        for triplet in data:
            row = next(reader)
            if len(triplet[0]) != 3 or all(v == 0 for v in triplet[0]):
                continue
            triplet_index += 1
            next_row = worksheet.max_row + 1
            human_answers = human_answer(row[17:28:2])
            image_urls = [row[0], row[5], row[10], human_answers[0], human_answers[1], human_answers[2]]
            agreement_val = max(human_answers[0], human_answers[1], human_answers[2])
            # Write the triplet index
            worksheet.cell(row=next_row, column=1, value=triplet_index)
            
            # Write the image URLs
            for col_idx, url in enumerate(image_urls, start=2):
                worksheet.cell(row=next_row, column=col_idx, value=url)

            max_correlation = 0
            max_column = 0
            for layer_num, coefficients in enumerate(triplet):
                worksheet.cell(row=worksheet.max_row + 1, column=8, value=f"Layer {layer_num}")
                # Initialize variables to track the correlations
                correlations = [coefficient[0] for coefficient in coefficients]
                max_correlation = max(correlations)
                max_column = correlations.index(max_correlation)
                # Calculate the mean of the two leftover correlations
                leftover_correlations = [corr for i, corr in enumerate(correlations) if i != max_column]
                dominance_val = max_correlation - sum(leftover_correlations) / len(leftover_correlations)

                for col_idx, coefficient in enumerate(coefficients, start=9):
                    worksheet.cell(row=worksheet.max_row, column=col_idx, value=coefficient[0])

                worksheet.cell(row=worksheet.max_row, column=col_idx + 1, value=human_answers[max_column])
                worksheet.cell(row=worksheet.max_row, column=col_idx + 2, value=dominance_val)
                worksheet.cell(row=worksheet.max_row, column=col_idx + 3, value=agreement_val)
                max_column = 0
                max_correlation = 0

    workbook.save(excel_file_path)

def human_answer(answers) -> list:
    ans_counter = [0, 0, 0]
    for ans in answers:
        if ans == '1':
            ans_counter[0] += 1
        elif ans == '2':
            ans_counter[1] += 1
        else:
            ans_counter[2] += 1
    return [ans_counter[0] / len(answers), ans_counter[1] / len(answers), ans_counter[2] / len(answers)]

# Usage
csv_file_path = r"C:\Users\ROEE\vit_data\fec\faceexp-comparison-data-test-public.csv"
excel_output_folder = r"C:\Users\ROEE\vit_face_expression\test"
image_folder_path = r"C:\\Users\\ROEE\\vit_data\\fec_dataset"
all_triplets_paths = get_folders_in_folder(image_folder_path)
print("should be a lot folders, is:", len(all_triplets_paths))
analyzed_images = []
correlation_coefficients = []
triplets_correlations = []

batch_size = 100  # Number of triplets to process in each batch
for batch_start in range(0, len(all_triplets_paths), batch_size):
    batch_end = min(batch_start + batch_size, len(all_triplets_paths))
    batch_triplets_paths = all_triplets_paths[batch_start:batch_end]
    
    for triplet_folder in batch_triplets_paths:
        print(triplet_folder)
        images_path_list = get_files_in_folder(triplet_folder)
        if len(images_path_list) < 3:
            continue
        for image_path in images_path_list:
            if image_is_valid(image_path):
                analyzed_image = run_model(image_path)
                analyzed_images.append(analyzed_image)
        if len(analyzed_images) == 3:
            for layer_num in range(13):
                correlation_coefficient = pearson_correlation(analyzed_images, layer_num)
                correlation_coefficients.append(correlation_coefficient)
            triplets_correlations.append(correlation_coefficients)
        analyzed_images = []
        correlation_coefficients = []

    write_to_excel_triplets_result(image_folder_path, triplets_correlations, excel_output_folder, csv_file_path)
    triplets_correlations = []  # Clear the list after writing to Excel
